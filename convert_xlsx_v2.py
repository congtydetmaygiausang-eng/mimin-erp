import openpyxl
import json
from datetime import datetime

src = "/workspace/attachments/49045a6b__e881bdf7-d7d5-4aa2-8250-882d7a69bdea.xlsx"
wb = openpyxl.load_workbook(src, data_only=True)

def date_to_str(d):
    if d is None: return ""
    if isinstance(d, datetime): return d.strftime("%d/%m/%Y")
    return str(d)

def s(v):
    if v is None: return ""
    return str(v).strip()

def sdt(v):
    if v is None: return ""
    if isinstance(v, float): return str(int(v))
    return str(v).strip()

# ====== SHEET 1: NHÂN SỰ (17 NV) ======
ws = wb["Nhan_su"]
ns_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    rec = {
        "stt": int(row[0]) if row[0] else 0,
        "maNV": s(row[1]),
        "hoTen": s(row[2]),
        "boPhan": s(row[3]),
        "chucVu": s(row[4]),
        "ngaySinh": date_to_str(row[5]),
        "gioiTinh": s(row[6]),
        "cccd": sdt(row[7]),
        "ngayCap": date_to_str(row[8]),
        "noiCap": s(row[9]),
        "sdt": sdt(row[10]),
        "email": s(row[11]),
        "diaChiTT": s(row[12]),
        "diaChiTamTru": s(row[13]),
        "viTri": s(row[14]),
        "ngayVaoLam": date_to_str(row[15]),
        "loaiHD": s(row[16]),
        "tinhTrangHN": s(row[17]),
        "soTK": s(row[18]),
        "nganHang": s(row[19]),
        "mst": s(row[20]),
        "bhxh": s(row[21]),
        "trangThai": s(row[22]) or s(row[23]),  # Try both
        "luongCB": int(row[24]) if isinstance(row[24], (int, float)) else 0,
        "loaiLuong": s(row[25]) if len(row) > 25 else "",
    }
    ns_rows.append(rec)

print(f"✅ Nhân sự: {len(ns_rows)} NV")
print(f"   - Có SĐT: {sum(1 for r in ns_rows if r['sdt'])}")
print(f"   - Có CCCD: {sum(1 for r in ns_rows if r['cccd'])}")
print(f"   - Có TK ngân hàng: {sum(1 for r in ns_rows if r['soTK'])}")
print(f"   - Có BHXH: {sum(1 for r in ns_rows if r['bhxh'])}")
print(f"   Bộ phận: {set(r['boPhan'] for r in ns_rows if r['boPhan'])}")
print(f"   Trạng thái: {set(r['trangThai'] for r in ns_rows if r['trangThai'])}")

# ====== SHEET 2: ĐỐI TÁC/NCC (35 đối tác) ======
ws = wb["Doi_tac_NCC"]
dt_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    rec = {
        "stt": int(row[0]) if row[0] else 0,
        "maDT": s(row[1]),
        "tenDonVi": s(row[2]),
        "nguoiLH": s(row[3]),
        "sdt": sdt(row[4]),
        "email": s(row[5]),
        "diaChi": s(row[6]),
        "boPhan": s(row[7]),
        "chucVu": s(row[8]),
        "soTK": s(row[9]),
        "nganHang": s(row[10]),
        "mst": s(row[11]),
        "loaiDT": s(row[12]),
        "trangThaiRaw": s(row[13]),
        "ghiChu": s(row[14]),
    }
    dt_rows.append(rec)

# Phân loại theo mã: GC-IN (in/thêu), GC-QUAN (may quần), GC-TRON (may tròn), GC-TRU (may trụ)
def get_stage(maDT):
    if "-IN-" in maDT: return "In / Thêu / Dập"
    if "-QUAN-" in maDT: return "May quần"
    if "-TRON-" in maDT: return "May áo tròn"
    if "-TRU-" in maDT: return "May áo trụ"
    return "Khác"

def get_status(s):
    s = (s or "").lower()
    if "ngưng" in s or "ngừng" in s: return "Ngưng"
    if "ít" in s: return "Ít làm"
    return "Đang hợp tác"

for r in dt_rows:
    r["congDoan"] = get_stage(r["maDT"])
    r["trangThai"] = get_status(r["trangThaiRaw"])

print(f"\n✅ Đối tác: {len(dt_rows)} đối tác")
print(f"   - In/Thêu: {sum(1 for r in dt_rows if r['congDoan']=='In / Thêu / Dập')}")
print(f"   - May quần: {sum(1 for r in dt_rows if r['congDoan']=='May quần')}")
print(f"   - May tròn: {sum(1 for r in dt_rows if r['congDoan']=='May áo tròn')}")
print(f"   - May trụ: {sum(1 for r in dt_rows if r['congDoan']=='May áo trụ')}")
print(f"   Trạng thái: Đang hợp tác={sum(1 for r in dt_rows if r['trangThai']=='Đang hợp tác')}, Ngưng={sum(1 for r in dt_rows if r['trangThai']=='Ngưng')}")
print(f"   Có SĐT: {sum(1 for r in dt_rows if r['sdt'])}")
print(f"   Có MST: {sum(1 for r in dt_rows if r['mst'])}")

# ====== SHEET 3: DATA SETUP ======
ws = wb["Data_Setup"]
bo_phan, chuc_vu = [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]: bo_phan.append(s(row[0]))
    if row[1]: chuc_vu.append(s(row[1]))

print(f"\n✅ Bộ phận: {len(bo_phan)} → {bo_phan}")
print(f"✅ Chức vụ: {len(chuc_vu)} → {chuc_vu}")

# ====== EXPORT JSON ======
out = {
    "nhanSu": ns_rows,
    "doiTac": dt_rows,
    "boPhan": bo_phan,
    "chucVu": chuc_vu,
}
with open("/workspace/mimin-erp/data-v2.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"\n📄 Đã lưu JSON: /workspace/mimin-erp/data-v2.json")

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# ======================= 1. NHÂN VIÊN LƯƠNG CỨNG =======================
data_luong_cung = [
    ["STT", "Họ tên", "Việc chính", "Lương (VNĐ/tháng)"],
    [8, "BÙI THỊ THANH", "Kế toán điều phối SX", 8000000],
    [9, "ĐỖ THỊ HUYỀN", "QL Khách hàng Sỉ", 7000000],
    [10, "NGUYỄN NGỌC CẨM VY", "Content - Media", 8000000],
    [11, "HUỲNH XUÂN HÒA", "Chưa phân công", 10000000],
    [12, "NGUYỄN QUỐC HẬU", "Nhân viên Kho", 7000000],
    [13, "NGUYỄN THỊ HOÀNG OANH", "Chưa phân công", 12000000],
]

# ======================= 2. NHÂN VIÊN LƯƠNG SẢN PHẨM =======================
data_luong_sp = [
    ["STT", "Họ tên", "Việc chính", "Đơn giá (đ/cái)", "Ghi chú chi tiết"],
    [1, "Nguyễn Văn Ruộng", "Khuy nút", 750, "Áp dụng chung"],
    [2, "Nguyễn Minh Đức", "Ủi", "800/700/600", "Áo trụ: 800, Áo tròn: 700, Quần: 600"],
    [3, "Lê Định", "Ủi", "800/700/600", "Áo trụ: 800, Áo tròn: 700, Quần: 600"],
    [4, "Trương Minh Tâm", "Ủi", "800/700/600", "Áo trụ: 800, Áo tròn: 700, Quần: 600"],
    [5, "Nguyễn Thị Mỹ Nhi", "Gấp xếp", "1300/800/1500/1000", "Bộ thường: 1300, Áo thường: 800, Bộ trắng: 1500, Áo trắng: 1000"],
    [6, "Võ Thị Mỹ Phương", "Gấp xếp", "1300/800/1500/1000", "Bộ thường: 1300, Áo thường: 800, Bộ trắng: 1500, Áo trắng: 1000"],
    [7, "Nguyễn Thị Bé", "Gấp xếp", "1300/800/1500/1000", "Bộ thường: 1300, Áo thường: 800, Bộ trắng: 1500, Áo trắng: 1000"],
    [14, "PHẠM VĂN ĐỆ", "Cắt", "1400/1200/900", "Áo trụ: 1400, Áo tròn: 1200, Quần: 900"],
    [15, "DƯƠNG TẤN VĨNH", "Cắt", "1400/1200/900", "Áo trụ: 1400, Áo tròn: 1200, Quần: 900"],
    [16, "NGUYỄN QUỐC MINH", "Cắt", "1400/1200/900", "Áo trụ: 1400, Áo tròn: 1200, Quần: 900"],
    [17, "TRƯƠNG VĂN NHẪN", "Cắt", "1400/1200/900", "Áo trụ: 1400, Áo tròn: 1200, Quần: 900"],
]

# ======================= 3. GIA CÔNG NGOÀI (35 đối tác) =======================
data_gia_cong = [
    ["STT", "Tên thợ / Chủ xưởng", "Công đoạn nhận", "Mã NCC", "Trạng thái", "Đơn giá ghi nhớ"],
    [1, "Tiến Đạt", "In", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [2, "Bảo Ngân", "In, Dập", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [3, "Thanh Sơn", "In, Dập", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [4, "Dung", "May quần", "NCC_15", "Đang hợp tác", "10.000 – 11.000 đ/cái"],
    [5, "Minh Vy", "May quần", "NCC_16", "Đang hợp tác", "10.000 – 11.000 đ/cái"],
    [6, "Hương", "May quần", "NCC_18", "Đang hợp tác", "10.000 – 11.000 đ/cái"],
    [7, "Thơ", "May quần", "NCC_17", "Đang hợp tác", "10.000 – 11.000 đ/cái"],
    [8, "Trai", "May áo tròn", "NCC_10", "Đang hợp tác", "Trơn 7.500, Dây 8.500, Lé 9.500"],
    [9, "Thuận", "May áo tròn", "NCC_14", "Đang hợp tác", "Trơn 7.500, Dây 8.500, Lé 9.500"],
    [10, "Hằng", "May áo tròn", "NCC_12", "Đang hợp tác", "Trơn 7.500, Dây 8.500, Lé 9.500"],
    [11, "Chiến", "May áo tròn", "NCC_13", "Đang hợp tác", "Trơn 7.500, Dây 8.500, Lé 9.500"],
    [12, "Thông", "May áo trụ", "NCC_07", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [13, "Cúc", "May áo trụ", "NCC_08", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [14, "Liễu", "May áo trụ", "NCC_03", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [15, "Duẩn", "May áo trụ", "NCC_05", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [16, "Sản", "May áo trụ", "NCC_09", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [17, "Tý Sơn", "May áo trụ", "NCC_04", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [18, "Toàn", "May áo trụ", "NCC_06", "Đang hợp tác", "Trơn 15k, Dây 16k, Lé 17k"],
    [19, "Hạnh", "Thêu", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [20, "Vui", "Thêu", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [21, "Trung", "Chưa ghi", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [22, "Quang", "Chưa ghi", "", "Đang hợp tác", "CHƯA CÓ GIÁ"],
    [23, "Ánh", "Chưa ghi", "", "Ít làm", "CHƯA CÓ GIÁ"],
    [24, "Bình", "Chưa ghi", "", "Ít làm", "CHƯA CÓ GIÁ"],
    [25, "Hiền", "Chưa ghi", "", "Ít làm", "CHƯA CÓ GIÁ"],
    [26, "Toản", "Chưa ghi", "", "Ngưng", "CHƯA CÓ GIÁ"],
    [27, "Tuấn", "May áo tròn", "NCC_11", "Ngưng", "Đã ngưng"],
    [28, "Phong", "Chưa ghi", "", "Ngưng", ""],
    [29, "Phúc", "Chưa ghi", "", "Ngưng", ""],
    [30, "Trí", "Chưa ghi", "", "Ngưng", ""],
    [31, "Mộng", "Chưa ghi", "", "Ngưng", ""],
    [32, "Thắng", "Chưa ghi", "", "Ngưng", ""],
    [33, "Kiếm", "Chưa ghi", "", "Ngưng", ""],
    [34, "Thiện", "Chưa ghi", "", "Ngưng", ""],
    [35, "Kiên", "Chưa ghi", "", "Ngưng", ""],
]

# ======================= 4. NHÀ CUNG CẤP (16 NCC) =======================
data_ncc = [
    ["STT", "Tên nhà cung cấp", "Vai trò / Dịch vụ", "Đơn giá (nếu có)", "Công nợ hiện tại (VNĐ)"],
    [1, "Công ty Lucky Avanti", "Bán sợi", "", 0],
    [2, "Công ty TNHH Thương mại Quốc tế Sammoon", "Bán sợi", "", 909052000],
    [3, "Công ty TNHH Sản xuất Thương mại Dệt May Hải Dương", "Dệt", "10.000 đ/kg", 183944000],
    [4, "Công ty TNHH Một thành viên Dệt Nhuộm Thái Thành", "Nhuộm", "", 0],
    [5, "Công ty Cổ phần Dệt Nhuộm Phú Long", "Bo cổ", "", 184369120],
    [6, "Hộ kinh doanh Vũ Văn Hiệp", "Bo cổ", "", 74315000],
    [7, "Công ty TNHH Phụ liệu May mặc Tường Vy", "Thun & Giấy gấp xếp", "", 10200000],
    [8, "Công ty TNHH Thương mại Dịch vụ Hằng Lữ", "Dây kéo", "", 91500000],
    [9, "(chưa có tên) — Dây xỏ mạc", "Dây xỏ mạc", "", 0],
    [10, "(chưa có tên) — Dây luồn quần", "Dây luồn quần", "", 0],
    [11, "Công ty TNHH Sản xuất và Thương mại Nhãn mác Hải Nam", "Nhãn thẻ bài", "", 0],
    [12, "Công ty TNHH In ấn Thông Anh", "Nhãn size", "", 21047904],
    [13, "Công ty TNHH Sản xuất Kinh doanh Thương mại Bao bì Đại Hoàng Phúc", "Túi zip", "", 0],
    [14, "CÔNG TY GIGATEX = Dệt Nhuộm Thái Thành (anh Hùng)", "Nhuộm", "", 639347450],
    [15, "CÔNG TY TNHH DỆT BO HẢI ÂU", "Bo cổ", "", 31795000],
    [16, "CÔNG TY TNHH BAO BÌ PHÚC VINH", "Khác", "", 3450000],
]

# ======================= TẠO FILE EXCEL =======================
wb = Workbook()
wb.remove(wb.active)

# Sheet 1: Lương cứng
ws1 = wb.create_sheet("1. Lương cứng")
df1 = pd.DataFrame(data_luong_cung[1:], columns=data_luong_cung[0])
for r in dataframe_to_rows(df1, index=False, header=True):
    ws1.append(r)

# Sheet 2: Lương sản phẩm
ws2 = wb.create_sheet("2. Lương sản phẩm")
df2 = pd.DataFrame(data_luong_sp[1:], columns=data_luong_sp[0])
for r in dataframe_to_rows(df2, index=False, header=True):
    ws2.append(r)

# Sheet 3: Gia công ngoài
ws3 = wb.create_sheet("3. Gia công ngoài")
df3 = pd.DataFrame(data_gia_cong[1:], columns=data_gia_cong[0])
for r in dataframe_to_rows(df3, index=False, header=True):
    ws3.append(r)

# Sheet 4: Nhà cung cấp
ws4 = wb.create_sheet("4. Nhà cung cấp")
df4 = pd.DataFrame(data_ncc[1:], columns=data_ncc[0])
for r in dataframe_to_rows(df4, index=False, header=True):
    ws4.append(r)

# Định dạng độ rộng cột
for ws in [ws1, ws2, ws3, ws4]:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width

# Lưu file
output_path = "/workspace/mimin-erp/DANH_SACH_TONG_HOP.xlsx"
wb.save(output_path)

print(f"✅ Đã xuất file thành công: {output_path}")
print("   Gồm 4 sheet: Lương cứng, Lương sản phẩm, Gia công ngoài, Nhà cung cấp")

# Tổng kết
print("\n📊 Tổng kết data:")
print(f"   - Lương cứng: {len(data_luong_cung)-1} nhân viên")
print(f"   - Lương sản phẩm: {len(data_luong_sp)-1} nhân viên")
print(f"   - Gia công ngoài: {len(data_gia_cong)-1} đối tác")
print(f"   - Nhà cung cấp: {len(data_ncc)-1} NCC")

# Tính tổng lương cứng
tong_luong_cung = sum(row[3] for row in data_luong_cung[1:])
print(f"   - Tổng lương cứng/tháng: {tong_luong_cung:,} đ")

# Tính tổng công nợ
tong_cong_no = sum(row[4] for row in data_ncc[1:])
print(f"   - Tổng công nợ NCC: {tong_cong_no:,} đ")
